from openpyxl import load_workbook


class Utils:
    def read_data_from_excel(file_name, sheet):
        data_list = []
        wb = load_workbook(filename=file_name)
        sh = wb[sheet]
        max_row = sh.max_row
        max_col = sh.max_column
        for i in range(2, max_row + 1):
            row = []
            for j in range(1, max_col + 1):
                row.append(sh.cell(row=i, column=j).value)
            data_list.append(row)

        return data_list